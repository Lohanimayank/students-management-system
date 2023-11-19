from tkinter import *
import tkinter as tk
from tkinter import ttk
from datetime import date
from tkinter import filedialog 
from tkinter.ttk import Combobox
from tkinter import messagebox
from PIL import Image,ImageTk
import os
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

from tkinter .messagebox import showinfo


background="#06283D"
framebg="#EDEDED"
framefg="#06283D"


root=tk.Tk()
root.title("student registration")
root.geometry('1250x700+210+100')
root.config(bg=background)




file=pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"

    file.save('Student_data.xlsx')

#EXIT
def Exit():
    root.destroy()



#for showing image
def showimage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                       title="select image file",filetype=(("JPG File","*.jpg"),
                                                                           ("PNG File","*.png"),
                                                                           ("All files","*.txt")))
    img=(Image.open(filename))
    resized_image= img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


#Registration No.
#automatic registration entry system

def registration_no():
   file=openpyxl.load_workbook('Student_data.xlsx')
   sheet=file.active
   row=sheet.max_row

   max_row_value=sheet.cell(row=row,column=1).value

   try:
       Registration.set(max_row_value+1)

   except:
      Registration.set("1")
      
      

#clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')
    F_Name.set('')
    M_Name.set('')
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Section.set("Select")

    registration_no()


    saveButton.config(state ='normal')


    img1=PhotoImage(file='Images/—Pngtree—reload user icon_4893685.png')
    lbl.config(image=img1)
    lbl.image=img1

    img=""



#save
def Save():
 R1=Registration.get()
 N1=Name.get()
 C1=Section.get()
 try:
     G1=gender
 except:
     messagebox.showerror("error","Select Gender !")


 D2=DOB.get()
 D1=Date.get()
 Re1=Religion.get()
 S1=Skill.get()
 fathername=F_Name.get()
 mothername=M_Name.get()
 F1=Father_Occupation.get()
 M1=Mother_Occupation.get()

 if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or F1=="" or M1=="":
     messagebox.showerror("error","Few Data is missing!")

 else:
     file=openpyxl.load_workbook('Student_data.xlsx')
     sheet=file.active
     sheet.cell(column=1,row=sheet.max_row+1,value=R1)
     sheet.cell(column=2,row=sheet.max_row,value=N1)
     sheet.cell(column=3,row=sheet.max_row,value=C1)
     sheet.cell(column=4,row=sheet.max_row,value=G1)
     sheet.cell(column=5,row=sheet.max_row,value=D2)
     sheet.cell(column=6,row=sheet.max_row,value=D1)
     sheet.cell(column=7,row=sheet.max_row,value=Re1)
     sheet.cell(column=8,row=sheet.max_row,value=S1)
     sheet.cell(column=9,row=sheet.max_row,value=fathername)
     sheet.cell(column=10,row=sheet.max_row,value=mothername)
     sheet.cell(column=11,row=sheet.max_row,value=F1)
     sheet.cell(column=12,row=sheet.max_row,value=M1)

     file.save(r'Student_data.xlsx')

     try:
         img.save("Student_Image/"+str(R1)+".jpg")
     except:
         messagebox.showinfo("info","Profile Picutre is not available!!!")

     messagebox.showinfo("info","Sucessfully data entered!!!")

     Clear()

     registration_no()
     

 

#search

def search():
    text=Search.get()

    Clear()
    saveButton.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
#          print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

#  print(reg_no_position)
# print(reg_number)    

    try:
      print(str(name))
    except:
      messagebox.showerror("Invalid","Invalid registration number!!!")


# rge_number from A1 to 1
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value

    Registration.set(x1)
    Name.set(x2)
    Section.set(x3)

    if x4=='female':
        R2.select()
    elif x4=='Male':
        R1.select()
    else:
        R3.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    img=(Image.open("Student_Image/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2


#update
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Section.get()
 
    selection()
    G1=gender   
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skill.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]

            print(reg_number)

    #sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Re1)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=fathername)
    sheet.cell(column=10,row=int(reg_number),value=mothername)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)

    file.save(r'Student_data.xlsx')

    try:
        img.save("Student_Image/"+str(R1)+".jpg")
   
   
    except:
        pass

    messagebox.showinfo("Update","Update successfully")

    Clear()



    

#for gender 
def selection():
    global gender
    value=radio.get()
    if value==1:
          gender="Male"
          
    elif value==2:
          gender="Female"
          
    else:
        gender="others"  
        
#top frames
Label(root,text="Email:mayanksproject@gmail.com",width=10,height=3,bg="#8B8B83",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=3,bg="#808080",fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)

#search box to update
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=10,font="arial 20").place(x=960,y=80)
imageicon3=PhotoImage(file="Images/icons8-search-client-50.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=100,bg='#68ddfa',font="arial 13 bold",command=search)
Srch.place(x=1100,y=75)


imageicon4=PhotoImage(file="Images/reload.png")
Update_button=Button(root,image=imageicon4,bg='#c36464',command=Update)
Update_button.place(x=110,y=64)

#Registration and Date
Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=160)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=160)

Registration=IntVar()
Date=StringVar()

reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=160)

registration_no()

today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=160)

Date.set(d1)



#student details
obj=LabelFrame(root,text="student Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)


Label(obj,text="Section:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=50)


DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=100)


radio=IntVar()
R1=Radiobutton(obj,text="Male", variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

R2=Radiobutton(obj,text="Female", variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)


R3=Radiobutton(obj,text="others", variable=radio,value=3,bg=framebg,fg=framefg,command=selection)
R3.place(x=263,y=150)

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")
religion_entry.place(x=630,y=100)

Skill=StringVar()
skill_entry=Entry(obj,textvariable=Skill,width=20,font="arial 10")
skill_entry.place(x=630,y=150)



Section=Combobox(obj,value=['A','B','C','D','E','F','G','H','I','J','K','L',],font="Roboto 10",width=17,state="r")
Section.place(x=630,y=50)
Section.set("Select")


#parentdetails
obj2=LabelFrame(root,text="parent Detail",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)


Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

F_Name=StringVar()
f_entry=Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)


Father_Occupation=StringVar()
FO_entry=Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
FO_entry.place(x=160,y=100)


Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

M_Name=StringVar()
M_entry=Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
M_entry.place(x=630,y=50)


Mother_Occupation=StringVar()
MO_entry=Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
MO_entry.place(x=630,y=100)




#image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Images/—Pngtree—reload user icon_4893685.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)


#button

Button(root,text="upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="SAVE",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Save)
saveButton.place(x=1000,y=450)

Button(root,text="RESET",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Clear).place(x=1000,y=530)

Button(root,text="EXIT",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Exit).place(x=1000,y=610)






root.mainloop()
